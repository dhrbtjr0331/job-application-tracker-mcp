"""Excel spreadsheet handling for job application tracking"""

from datetime import datetime
from pathlib import Path
from typing import Dict, List, Any, Set

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from job_application_tracker.config.settings import Settings
from job_application_tracker.utils.logging_config import get_logger

logger = get_logger(__name__)

class ExcelTracker:
    """Handles Excel spreadsheet operations for job tracking"""
    
    def __init__(self, excel_path: str = None):
        self.excel_path = Path(excel_path or Settings.DEFAULT_EXCEL_PATH)
        self.sheet_name = Settings.DEFAULT_SHEET_NAME
        self.headers = Settings.EXCEL_HEADERS
    
    def update_applications(self, applications: List[Dict[str, Any]]) -> bool:
        """Update Excel sheet with application data"""
        try:
            wb, ws = self._get_or_create_workbook()
            self._ensure_headers(ws)
            
            existing_message_ids = self._get_existing_message_ids(ws)
            added_count = self._add_new_applications(ws, applications, existing_message_ids)
            
            wb.save(self.excel_path)
            logger.info(f"Excel updated: {added_count} new applications added")
            return True
            
        except Exception as e:
            logger.error(f"Error updating Excel: {e}")
            return False
    
    def get_application_summary(self) -> Dict[str, Any]:
        """Get summary of tracked applications"""
        try:
            if not self.excel_path.exists():
                return {"error": f"Excel file not found: {self.excel_path}"}
            
            wb = load_workbook(self.excel_path)
            if self.sheet_name not in wb.sheetnames:
                return {"error": f"No {self.sheet_name} sheet found in Excel file"}
            
            ws = wb[self.sheet_name]
            status_counts, total_applications = self._count_applications_by_status(ws)
            
            if total_applications == 0:
                return {"error": "No applications found in Excel file"}
            
            return {
                "total_applications": total_applications,
                "status_counts": status_counts,
                "file_path": str(self.excel_path)
            }
            
        except Exception as e:
            logger.error(f"Error reading Excel summary: {e}")
            return {"error": f"Error reading Excel file: {str(e)}"}
    
    def _get_or_create_workbook(self) -> tuple[Workbook, Worksheet]:
        """Get existing workbook or create new one"""
        if self.excel_path.exists():
            wb = load_workbook(self.excel_path)
            if self.sheet_name in wb.sheetnames:
                ws = wb[self.sheet_name]
            else:
                ws = wb.create_sheet(self.sheet_name)
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = self.sheet_name
        
        return wb, ws
    
    def _ensure_headers(self, ws: Worksheet) -> None:
        """Ensure headers are present in the worksheet"""
        if ws.max_row == 1 and ws['A1'].value is None:
            for col, header in enumerate(self.headers, 1):
                ws.cell(row=1, column=col, value=header)
    
    def _get_existing_message_ids(self, ws: Worksheet) -> Set[str]:
        """Get set of existing message IDs to avoid duplicates"""
        existing_message_ids = set()
        message_id_col = 7  # Message ID column
        
        for row in range(2, ws.max_row + 1):
            message_id = ws.cell(row=row, column=message_id_col).value
            if message_id:
                existing_message_ids.add(message_id)
        
        return existing_message_ids
    
    def _add_new_applications(self, ws: Worksheet, applications: List[Dict[str, Any]], 
                            existing_message_ids: Set[str]) -> int:
        """Add new applications to the worksheet"""
        added_count = 0
        
        for app in applications:
            if app['message_id'] not in existing_message_ids:
                row = ws.max_row + 1
                self._write_application_row(ws, row, app)
                added_count += 1
        
        return added_count
    
    def _write_application_row(self, ws: Worksheet, row: int, app: Dict[str, Any]) -> None:
        """Write application data to a specific row"""
        ws.cell(row=row, column=1, value=app['company'])
        ws.cell(row=row, column=2, value=app['position'])
        ws.cell(row=row, column=3, value=app['status'])
        ws.cell(row=row, column=4, value=app['date'])
        ws.cell(row=row, column=5, value=app['subject'])
        ws.cell(row=row, column=6, value=app['sender'])
        ws.cell(row=row, column=7, value=app['message_id'])
        ws.cell(row=row, column=8, value=datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
    
    def _count_applications_by_status(self, ws: Worksheet) -> tuple[Dict[str, int], int]:
        """Count applications by status"""
        status_counts = {}
        total_applications = 0
        status_col = 3  # Status column
        
        for row in range(2, ws.max_row + 1):
            status = ws.cell(row=row, column=status_col).value
            if status:
                total_applications += 1
                status_counts[status] = status_counts.get(status, 0) + 1
        
        return status_counts, total_applications